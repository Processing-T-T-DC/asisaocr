from openpyxl import load_workbook
from datetime import datetime

from constants import (
    EXCEL_INPUT,
    EXCEL_OUTPUT,
    SHEET_NAME
)


# =====================
# CONFIGURACIÓN COLUMNAS ENTRADA
# =====================
COL_TRATAMIENTO = 1   # A
COL_FINALIDAD   = 2   # B
COL_NOMBRE      = 3   # C
COL_DESCRIPCION = 4   # D


# =====================
# CABECERAS SALIDA
# =====================
HEADER_ESTADO   = "Estado Selenium"
HEADER_FICHERO  = "Fichero descargado"
HEADER_FECHA    = "Fecha ejecución"
HEADER_ERROR    = "Error"


# =====================
# CARGA EXCEL
# =====================
def cargar_excel():
    wb = load_workbook(EXCEL_INPUT)
    ws = wb[SHEET_NAME]
    return wb, ws


# =====================
# OBTENER / CREAR COLUMNA
# =====================
def obtener_columna(ws, header):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header
    return col


# =====================
# OBTENER COLUMNAS RESULTADO
# =====================
def obtener_columnas_resultado(ws):
    return {
        "estado": obtener_columna(ws, HEADER_ESTADO),
        "fichero": obtener_columna(ws, HEADER_FICHERO),
        "fecha": obtener_columna(ws, HEADER_FECHA),
        "error": obtener_columna(ws, HEADER_ERROR),
    }


# =====================
# ITERAR FILAS
# =====================
def iterar_filas(ws):
    filas = []

    for row in range(2, ws.max_row + 1):
        tratamiento = ws.cell(row=row, column=COL_TRATAMIENTO).value
        finalidad   = ws.cell(row=row, column=COL_FINALIDAD).value
        nombre      = ws.cell(row=row, column=COL_NOMBRE).value
        descripcion = ws.cell(row=row, column=COL_DESCRIPCION).value

        if not any([tratamiento, finalidad, nombre, descripcion]):
            continue

        filas.append((
            row,
            str(tratamiento).strip() if tratamiento else "",
            str(finalidad).strip() if finalidad else "",
            str(nombre).strip() if nombre else "",
            str(descripcion).strip() if descripcion else ""
        ))

    return filas


# =====================
# ESCRIBIR RESULTADO
# =====================
def escribir_resultado(ws, row, col, value):
    ws.cell(row=row, column=col).value = value


# =====================
# GUARDAR EXCEL
# =====================
def guardar_excel(wb):
    wb.save(EXCEL_OUTPUT)

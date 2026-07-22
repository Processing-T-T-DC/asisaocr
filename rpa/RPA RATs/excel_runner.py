from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

EXCEL_FILE = "RAT_resultado.xlsx"
SHEET_NAME = "Ejecuciones"

# =====================
# COLUMNAS LOG
# =====================
COL_ESTADO  = 1  # A
COL_SECCION = 2  # B
COL_POSICION = 3  # C
COL_FECHA   = 4  # D
COL_ERROR   = 5  # E

# =====================
# CABECERAS
# =====================
HEADERS = [
    "Estado Selenium",
    "Sección del fichero generado",
    "Fichero generado",
    "Fecha ejecución",
    "Error"
]

# =====================
# CARGA / CREA EXCEL
# =====================
def cargar_excel():

    if not os.path.exists(EXCEL_FILE):
        print("📄 Excel no existe, creando uno nuevo")

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        _escribir_cabeceras(ws)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        _escribir_cabeceras(ws)
    else:
        ws = wb[SHEET_NAME]

    return wb, ws


# =====================
# CABECERAS
# =====================
def _escribir_cabeceras(ws):
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col).value = header


# =====================
# SIGUIENTE FILA
# =====================
def _siguiente_fila(ws):
    return ws.max_row + 1


# =====================
# ESCRIBIR LOG
# =====================
def escribir_resultado(
    ws,
    estado,
    fichero=None,
    posicion=None,
    error=None
):
    row = _siguiente_fila(ws)

    ws.cell(row=row, column=COL_ESTADO).value = estado
    ws.cell(row=row, column=COL_SECCION).value = fichero
    ws.cell(row=row, column=COL_POSICION).value = posicion
    ws.cell(row=row, column=COL_FECHA).value = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    ws.cell(row=row, column=COL_ERROR).value = error


# =====================
# GUARDAR
# =====================
def guardar_excel(wb):
    wb.save(EXCEL_FILE)

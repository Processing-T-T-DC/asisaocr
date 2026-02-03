import os
from src.model.model import Model, WritableEntry, WritableFile, WritableParameters, Writer, WritingTarget
import openpyxl
from openpyxl.cell import Cell
from src.errors import Error, FileWriteError
from openpyxl.utils import get_column_letter


class ExcelWriter(Writer):

    class WritableExcelFile(WritableFile):

        class WritableExcelParameters(WritableParameters):
            row: int
            column: int

            def __init__(self, row: int, column: int):
                self.row = row
                self.column = column

        class WritableExcelEntry(WritableEntry):
            content: str
            parameters: "ExcelWriter.WritableExcelFile.WritableExcelParameters"

            def __init__(self, content: str, parameters: "ExcelWriter.WritableExcelFile.WritableExcelParameters"):
                self.content = content
                self.parameters = parameters

        entries: list[WritableExcelEntry] = []

        def __init__(self, target: str, entries: list[WritableExcelEntry]):
            self.entries = entries
            self.target = target

    def write(self, writable_file: WritableExcelFile, template: str | None = None) -> None | Error:
        """Write the parsed file to excel format and return as bytes."""
        # Dummy implementation for demonstration

        if template is not None:

            if template == "templates/RAT_Modelo de datos_template.xlsx" and "RAT_master.xlsx" in os.listdir("output/"):
                workbook = openpyxl.load_workbook("output/RAT_master.xlsx")
            
            elif template == "templates/AARR_Modelo de datos_template.xlsx" and  "evaluaciones_objetivas_master.xlsx" in os.listdir("output/"):
                workbook = openpyxl.load_workbook("output/evaluaciones_objetivas_master.xlsx")
            
            elif template == "templates/PIA_Modelo de datos_template.xlsx" and  "pia_master.xlsx" in os.listdir("output/"):
                workbook = openpyxl.load_workbook("output/pia_master.xlsx")
            else:
                workbook = openpyxl.load_workbook(template)
        else:
            workbook = openpyxl.Workbook()
        
        sheet = workbook.active

        if sheet is None:
            raise FileWriteError("Sheet is empty right after creation.")
        
        if template == "templates/RAT_Modelo de datos_template.xlsx" or template == "templates/AARR_Modelo de datos_template.xlsx" or template == "templates/PIA_Modelo de datos_template.xlsx":
            # check last filled row
            col = 2  # A = 1, B = 2, etc.
            last_row: int
            for row in range(sheet.max_row, 0, -1):
                if sheet.cell(row=row, column=col).value is not None:
                    last_row = row
                    break
            else:
                last_row = 0  # column is empty

            # fill row
            
            for entry in writable_file.entries:
                if entry.content is None:
                    continue

                value = entry.content
                sheet[f"{get_column_letter(entry.parameters.column)}{last_row + entry.parameters.row}"] = value
        else:
            for entry in writable_file.entries:
                value = entry.content
                sheet[f"{get_column_letter(entry.parameters.column)}{entry.parameters.row + 1}"] = value

        if self.target is None:
            raise FileWriteError("Writing target is None.")
            
        workbook.save(self.target)

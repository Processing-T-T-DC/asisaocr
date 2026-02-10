from typing import Literal, TypeVar, cast, overload

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from src.errors import ValidationError
from src.model.model import Field, File, Model, ParsedFile
from src.model.writers.excel_writer import ExcelWriter

type ColumnKey = Literal[
    "tratamiento",
    "responsable_del_tratamiento",
    "data_owner",
    "data_owner_externo",
    "delegado_de_proteccion_de_datos",
    "delegado_de_proteccion_de_datos_externo",
    "otras_figuras_o_responsables",
    "figura_en_que_se_actua_respecto_al_tratamiento",
    "nombre",
    "nombre_del_destinatario",
    "funcion_del_destinatario",
    "encargo_de_tratamiento",
    "observaciones_de_encargo_de_tratamiento",
    "politicas_de_conservacion",
    "resumen_de_politicas_de_conservacion",
    "observaciones_de_politicas_de_conservacion",
    "responsable_del_tratamiento",
    "data_owner",
    "data_owner_externo",
    "delegado_de_proteccion_de_datos",
    "delegado_de_proteccion_de_datos_externo",
    "otras_figuras_o_responsables",
    "figura_en_que_se_actua_respecto_al_tratamiento",
    "sistema_de_informacion",
    "resumen_medidas_de_seguridad",
    "se_tratan_datos_colectivos",
    "datos_colectivos",
    "se_tratan_datos_identificativos",
    "datos_de_caracter_identificativo",
    "se_tratan_datos_sensibles_o_especialmente_protegidos",
    "datos_especialmente_protegidos",
    "otros_datos_especialmente_protegidos",
    "se_tratan_datos_comision_de_infracciones_penales_o_administrativas",
    "datos_comision_de_infracciones_penales_o_administrativas",
    "se_tratan_datos_de_caracteristicas_personales",
    "datos_de_caracteristicas_personales",
    "se_tratan_datos_de_circunstancias_sociales",
    "circunstancias_sociales",
    "se_tratan_datos_academicos_y_profesionales",
    "datos_academicos_y_profesionales",
    "se_tratan_datos_de_detalle_de_empleo",
    "datos_de_detalle_de_empleo",
    "se_tratan_datos_de_informacion_comercial",
    "informacion_comercial",
    "se_tratan_datos_sobre_transacciones_de_bienes_y_servicios",
    "transacciones_de_bienes_y_servicios",
    "se_tratan_datos_economicos_financieros_y_de_seguros",
    "datos_economicos_financieros_y_de_seguros",
    "se_tratan_otras_categorias_de_datos",
    "otras_categoria_de_datos",
    "base_legitimadora_del_tratamiento",
    "ley_aplicable_o_mision_base_legitimadora_del_tratamiento",
    "descripcion_base_legitimadora_del_tratamiento",
    "identificacion_de_transferencia",
    "resumen_de_transferencias_internacionales",
    "categorias_de_datos",
    "garantias_de_transferencias_internacionales",
]

class RAT_Model(Model):

    FONT_SIZES_MAPPING = {
        18: "title", # Only title for both pdfs
        14: "H1", # 
        12: "H2", # 
        11: "H4", # 
    }

    @property
    def fields(self) -> list[Field]:
        return [
            Field("Tratamiento", "B1", "text", True),
            Field("Gestor de Datos", "C1", "text", False),
            Field("Delegado de Datos", "D1", "text", True),
            Field("Responsable del Tratamiento", "E1", "text", False),
            Field("Finalidad", "F1", "text", True),
            Field("¿Incluye datos colectivos?", "G1", "yes_no", True),
            Field("Detalle", "H1", "text", False),
            Field("¿Incluye datos identificativos o de contacto?", "I1", "yes_no", True),
            Field("Detalle", "J1", "text", False),
            Field("¿Incluye datos sensibles o especialmente protegidos?", "K1", "yes_no", True),
            Field("Datos especialmente protegidos", "L1", "text", False),
            Field("Otros Datos Especialmente protegidos", "M1", "text", False),
            Field("¿Incluye comisión de infracciones penales o administrativas?", "N1", "yes_no", True),
            Field("Detalle", "O1", "text", False),
            Field("¿Incluye datos de características personales?", "P1", "yes_no", True),
            Field("Detalle", "Q1", "text", False),
            Field("¿Incluye datos de circunstancias sociales?", "R1", "yes_no", True),
            Field("Detalle", "S1", "text", False),
            Field("¿Incluye datos académicos y profesionales?", "T1", "yes_no", True),
            Field("Detalle", "U1", "text", False),
            Field("¿Incluye datos de detalle del empleo?", "V1", "yes_no", True),
            Field("Detalle", "W1", "text", False),
            Field("¿Incluye datos de información comercial?", "X1", "yes_no", True),
            Field("Detalle", "Y1", "text", False),
            Field("¿Incluye datos de transacciones de bienes y servicios?", "Z1", "yes_no", True),
            Field("Detalle", "AA1", "text", False),
            Field("¿Incluye datos económicos, financieros y de seguros?", "AB1", "yes_no", True),
            Field("Detalle", "AC1", "text", False),
            Field("¿Incluye datos de otras categorías de datos?", "AD1", "yes_no", True),
            Field("Detalle", "AE1", "text", False),
            Field("Base legitimadora del tratamiento", "AF1", "text", False),
            Field("Descripción de la base legitmadora del tratamiento", "AG1", "text", False),
            Field("Encargo del tratamiento", "AH1", "text", False),
            Field("Destinatario", "AI1", "text", False),
            Field("Función del Destinatario", "AJ1", "text", False),
            Field("Conservación y Supresión - Resumen", "AK1", "text", False),
            Field("Conservación y Supresión - Observaciones", "AL1", "text", False),
            Field("Transferencias Internacionales - Identificación de la transferencia", "AM1", "text", False),
            Field("Transferencias Internacionales - Categorías de Datos", "AN1", "text", False),
            Field("Transferencias Internacionales - Resumen", "AO1", "text", False),
            Field("Transferencias Internacionales - Garantías", "AP1", "text", False),
            Field("Sistemas de Información", "AQ1", "text", False),
        ]


    output_mapping: dict[ColumnKey, str] = {
        "tratamiento": "B",
        "responsable_del_tratamiento": "E",
        "data_owner": "C",
        "data_owner_externo": "",
        "delegado_de_proteccion_de_datos": "D",
        "delegado_de_proteccion_de_datos_externo": "",
        "otras_figuras_o_responsables": "",
        "figura_en_que_se_actua_respecto_al_tratamiento": "",
        "nombre": "F",
        "nombre_del_destinatario": "AI",
        "funcion_del_destinatario": "AJ",
        "encargo_de_tratamiento": "AH",
        "observaciones_de_encargo_de_tratamiento": "AL",
        "politicas_de_conservacion": "",
        "resumen_de_politicas_de_conservacion": "AK",
        "observaciones_de_politicas_de_conservacion": "AL",
        # "responsable_del_tratamiento": "",
        # "data_owner": "",
        # "data_owner_externo": "",
        # "delegado_de_proteccion_de_datos": "",
        # "delegado_de_proteccion_de_datos_externo": "",
        # "otras_figuras_o_responsables": "",
        # "figura_en_que_se_actua_respecto_al_tratamiento": "",
        "sistema_de_informacion": "AQ",
        "resumen_medidas_de_seguridad": "",
        "se_tratan_datos_colectivos": "G",
        "datos_colectivos": "H",
        "se_tratan_datos_identificativos": "I",
        "datos_de_caracter_identificativo": "J",
        "se_tratan_datos_sensibles_o_especialmente_protegidos": "K",
        "datos_especialmente_protegidos": "L",
        "otros_datos_especialmente_protegidos": "M",
        "se_tratan_datos_comision_de_infracciones_penales_o_administrativas": "N",
        "datos_comision_de_infracciones_penales_o_administrativas": "O",
        "se_tratan_datos_de_caracteristicas_personales": "P",
        "datos_de_caracteristicas_personales": "Q",
        "se_tratan_datos_de_circunstancias_sociales": "R",
        "circunstancias_sociales": "S",
        "se_tratan_datos_academicos_y_profesionales": "T",
        "datos_academicos_y_profesionales": "U",
        "se_tratan_datos_de_detalle_de_empleo": "V",
        "datos_de_detalle_de_empleo": "W",
        "se_tratan_datos_de_informacion_comercial": "X",
        "informacion_comercial": "Y",
        "se_tratan_datos_sobre_transacciones_de_bienes_y_servicios": "Z",
        "transacciones_de_bienes_y_servicios": "AA",
        "se_tratan_datos_economicos_financieros_y_de_seguros": "AB",
        "datos_economicos_financieros_y_de_seguros": "AC",
        "se_tratan_otras_categorias_de_datos": "AD",
        "otras_categoria_de_datos": "AE",
        "base_legitimadora_del_tratamiento": "AF",
        "ley_aplicable_o_mision_base_legitimadora_del_tratamiento": "",
        "descripcion_base_legitimadora_del_tratamiento": "AG",
        "identificacion_de_transferencia": "AM",
        "resumen_de_transferencias_internacionales": "AO",
        "categorias_de_datos": "AN",
        "garantias_de_transferencias_internacionales": "AP"
    }
        

    class Responsables:
        responsable_del_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        data_owner: ExcelWriter.WritableExcelFile.WritableExcelEntry
        data_owner_externo: ExcelWriter.WritableExcelFile.WritableExcelEntry
        delegado_de_proteccion_de_datos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        delegado_de_proteccion_de_datos_externo: ExcelWriter.WritableExcelFile.WritableExcelEntry
        otras_figuras_o_responsables: ExcelWriter.WritableExcelFile.WritableExcelEntry
        figura_en_que_se_actua_respecto_al_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry

    class Finalidad:
        nombre: ExcelWriter.WritableExcelFile.WritableExcelEntry
        tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        
        nombre_del_destinatario: ExcelWriter.WritableExcelFile.WritableExcelEntry
        funcion_del_destinatario: ExcelWriter.WritableExcelFile.WritableExcelEntry
        
        encargo_de_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        observaciones_de_encargo_de_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        politicas_de_conservacion: ExcelWriter.WritableExcelFile.WritableExcelEntry
        resumen_de_politicas_de_conservacion: ExcelWriter.WritableExcelFile.WritableExcelEntry
        observaciones_de_politicas_de_conservacion: ExcelWriter.WritableExcelFile.WritableExcelEntry
        responsable_del_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        data_owner: ExcelWriter.WritableExcelFile.WritableExcelEntry
        data_owner_externo: ExcelWriter.WritableExcelFile.WritableExcelEntry
        delegado_de_proteccion_de_datos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        delegado_de_proteccion_de_datos_externo: ExcelWriter.WritableExcelFile.WritableExcelEntry
        otras_figuras_o_responsables: ExcelWriter.WritableExcelFile.WritableExcelEntry
        figura_en_que_se_actua_respecto_al_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        
        sistema_de_informacion: ExcelWriter.WritableExcelFile.WritableExcelEntry
        resumen_medidas_de_seguridad: ExcelWriter.WritableExcelFile.WritableExcelEntry
        se_tratan_datos_colectivos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_colectivos: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_identificativos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_de_caracter_identificativo: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_sensibles_o_especialmente_protegidos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_especialmente_protegidos: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        otros_datos_especialmente_protegidos: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_comision_de_infracciones_penales_o_administrativas: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_comision_de_infracciones_penales_o_administrativas: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_de_caracteristicas_personales: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_de_caracteristicas_personales: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        se_tratan_datos_de_circunstancias_sociales: ExcelWriter.WritableExcelFile.WritableExcelEntry
        circunstancias_sociales: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_academicos_y_profesionales: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_academicos_y_profesionales: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]
        
        se_tratan_datos_de_detalle_de_empleo: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_de_detalle_de_empleo: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        se_tratan_datos_de_informacion_comercial: ExcelWriter.WritableExcelFile.WritableExcelEntry
        informacion_comercial: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        se_tratan_datos_sobre_transacciones_de_bienes_y_servicios: ExcelWriter.WritableExcelFile.WritableExcelEntry
        transacciones_de_bienes_y_servicios: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        se_tratan_datos_economicos_financieros_y_de_seguros: ExcelWriter.WritableExcelFile.WritableExcelEntry
        datos_economicos_financieros_y_de_seguros: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        se_tratan_otras_categorias_de_datos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        otras_categoria_de_datos: list[ExcelWriter.WritableExcelFile.WritableExcelEntry]

        base_legitimadora_del_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        ley_aplicable_o_mision_base_legitimadora_del_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry
        descripcion_base_legitimadora_del_tratamiento: ExcelWriter.WritableExcelFile.WritableExcelEntry

        identificacion_de_transferencia: ExcelWriter.WritableExcelFile.WritableExcelEntry
        resumen_de_transferencias_internacionales: ExcelWriter.WritableExcelFile.WritableExcelEntry
        categorias_de_datos: ExcelWriter.WritableExcelFile.WritableExcelEntry
        garantias_de_transferencias_internacionales: ExcelWriter.WritableExcelFile.WritableExcelEntry

    responsables: Responsables | None
    finalidades: list[Finalidad]
    target: str | None
    output_row_start = 1

    _current_cell: Cell | ReadOnlyCell

    def _get_text_from_cell_or_empty_string(self, sheet: Worksheet, address: str) -> str:
        return cast(str, cast(Cell, sheet[address]).value or "")
    
    type WritableExcelParameters = ExcelWriter.WritableExcelFile.WritableExcelParameters
    
    T = TypeVar('T')

    def validate_output_for_file(self, file: File) -> bool | ValidationError:
        
        # Read file with 



        return False

    
    @overload
    def _get_list_of_vertically_aligned_rows_with_data(
        self,
        sheet: Worksheet,
        row: int,
        column: str,
        finalidad_height: int,
        return_type: Literal["str"],
    ) -> list[str]: ...

    @overload
    def _get_list_of_vertically_aligned_rows_with_data(
        self,
        sheet: Worksheet,
        row: int,
        column: str,
        finalidad_height: int,
        return_type: Literal["cell"],
    ) -> list[Cell]: ...


    def _get_list_of_vertically_aligned_rows_with_data(
        self,
        sheet: Worksheet,
        row: int,
        column: str,
        finalidad_height: int,
        return_type: Literal["str", "cell"],
    ):
        last_row = finalidad_height if finalidad_height != -1 else sheet.max_row
        rows = []

        current_row = row
        while True:
            for r in range(current_row, last_row + 1):
                cell = sheet[f"{column}{r}"]
                if cell.value is None:
                    current_row = r
                    break
                rows.append(cell.value if return_type == "str" else cell)
            else:
                break

            if current_row >= last_row:
                break
            current_row += 1

        return rows
    
    def _get_height_of_finalidad(self, sheet: Worksheet, row: int) -> int:
        current_row = row + 1
        for current_row in range(row+1, sheet.max_row + 1):
            if sheet[f"A{current_row}"].value is not None:
                return current_row - 1

        if current_row >= sheet.max_row:
            return -1
            

        return current_row

    def _get_next_column(self) -> str:
        next_cell = self._current_sheet.cell(self._current_cell.row, self._current_cell.column + 1)
        assert isinstance(next_cell, (Cell))

        self._current_cell = next_cell

        return next_cell.column_letter
    
    def _get_entry_from_cell(self, cell: Cell) -> ExcelWriter.WritableExcelFile.WritableExcelEntry:
        params = ExcelWriter.WritableExcelFile.WritableExcelParameters(cell.row, cell.column)
        return ExcelWriter.WritableExcelFile.WritableExcelEntry(self._get_text_from_cell_or_empty_string(self._current_sheet, cell.coordinate), params)

    def process(self, workbook: Workbook, target: str):
        self.finalidades = []
        
        for sheet in workbook.worksheets:
            self._current_sheet = sheet

            self.responsables = RAT_Model.Responsables()    

            self.responsables.responsable_del_tratamiento = self._get_entry_from_cell(sheet["A5"])
            self.responsables.data_owner = self._get_entry_from_cell(sheet["B5"])
            self.responsables.data_owner_externo = self._get_entry_from_cell(sheet["C5"])
            self.responsables.delegado_de_proteccion_de_datos = self._get_entry_from_cell(sheet["D5"]) 
            self.responsables.delegado_de_proteccion_de_datos_externo = self._get_entry_from_cell(sheet["E5"])
            self.responsables.otras_figuras_o_responsables = self._get_entry_from_cell(sheet["F5"])
            self.responsables.figura_en_que_se_actua_respecto_al_tratamiento = self._get_entry_from_cell(sheet["G5"])


            # find most immediate row below the finalidad field
            is_last_entry = False

            current_row = 11
            while not is_last_entry:
                starting_row = current_row
                for current_row in range(starting_row, sheet.max_row + 1):
                    if sheet[f"A{current_row}"].value is not None:
                        break

                if current_row >= sheet.max_row:
                    is_last_entry = True
                
                current_cell: Cell = sheet[f"A{current_row}"]
                self._current_cell = current_cell
                
                if current_cell.fill.bgColor.rgb == 'FF40A04C': # This is the greem color headers like "Finalidades no gestionadas" have
                    break

                finalidad_height = self._get_height_of_finalidad(sheet, current_row)


                
                finalidad = RAT_Model.Finalidad()
                finalidad.tratamiento = self._get_entry_from_cell(sheet["A3"])
                finalidad.tratamiento.content = finalidad.tratamiento.content.removeprefix("Responsable del tratamiento: ")
                finalidad.nombre = self._get_entry_from_cell(sheet[f"{self._current_cell.column_letter}{current_row}"])
                finalidad.nombre_del_destinatario = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.funcion_del_destinatario = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])

                finalidad.encargo_de_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.observaciones_de_encargo_de_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.politicas_de_conservacion = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.resumen_de_politicas_de_conservacion = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.observaciones_de_politicas_de_conservacion = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])

                # For some reason there are 7 columns that are always empty, and correspond to Finalidad. Let's skip them as 
                # that info was already obtained in the top side of the file 
                self._current_cell = self._current_cell.offset(0, 7)

                # finalidad.responsable_del_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.data_owner = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.data_owner_externo = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.delegado_de_proteccion_de_datos = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.delegado_de_proteccion_de_datos_externo = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.otras_figuras_o_responsables = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                # finalidad.figura_en_que_se_actua_respecto_al_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])

                finalidad.sistema_de_informacion = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.resumen_medidas_de_seguridad = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_colectivos = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                if (len(finalidad.datos_colectivos) > 0):
                    column = finalidad.datos_colectivos[0].parameters.column
                    
                    finalidad.se_tratan_datos_colectivos = ExcelWriter.WritableExcelFile.WritableExcelEntry("SI", ExcelWriter.WritableExcelFile.WritableExcelParameters(current_row, column))
                else:
                    finalidad.se_tratan_datos_colectivos = ExcelWriter.WritableExcelFile.WritableExcelEntry("NO", ExcelWriter.WritableExcelFile.WritableExcelParameters(current_row, self._current_cell.column))
                
                # finalidad.se_tratan_datos_identificativos = sheet[f"{self._get_next_column()}{current_row}"]
                # finalidad.datos_de_caracter_identificativo = self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, "T", finalidad_height)

                finalidad.se_tratan_datos_identificativos = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_de_caracter_identificativo = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_datos_sensibles_o_especialmente_protegidos = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_especialmente_protegidos = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                finalidad.otros_datos_especialmente_protegidos = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                
                finalidad.se_tratan_datos_comision_de_infracciones_penales_o_administrativas = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_comision_de_infracciones_penales_o_administrativas = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                
                finalidad.se_tratan_datos_de_caracteristicas_personales = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_de_caracteristicas_personales = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_datos_de_circunstancias_sociales = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.circunstancias_sociales = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                
                finalidad.se_tratan_datos_academicos_y_profesionales = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_academicos_y_profesionales = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]
                
                finalidad.se_tratan_datos_de_detalle_de_empleo = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_de_detalle_de_empleo = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_datos_de_informacion_comercial = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.informacion_comercial = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_datos_sobre_transacciones_de_bienes_y_servicios = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.transacciones_de_bienes_y_servicios = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_datos_economicos_financieros_y_de_seguros = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.datos_economicos_financieros_y_de_seguros = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.se_tratan_otras_categorias_de_datos = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.otras_categoria_de_datos = [self._get_entry_from_cell(cell) for cell in self._get_list_of_vertically_aligned_rows_with_data(sheet, current_row, self._get_next_column(), finalidad_height, "cell")]

                finalidad.base_legitimadora_del_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.ley_aplicable_o_mision_base_legitimadora_del_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.descripcion_base_legitimadora_del_tratamiento = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])

                finalidad.identificacion_de_transferencia = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.resumen_de_transferencias_internacionales = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.categorias_de_datos = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])
                finalidad.garantias_de_transferencias_internacionales = self._get_entry_from_cell(sheet[f"{self._get_next_column()}{current_row}"])

                self.finalidades.append(finalidad)

                current_row += 1
            
        
        self.target = target

    def create_writable_file(self, parsed_file: ParsedFile) -> ExcelWriter.WritableExcelFile:
        entries = []

        for name, entry in vars(self.responsables).items():

            for finalidad_index in range(len(self.finalidades)):
                name = cast(ColumnKey, name)
                mapped_output_column = self.output_mapping[name]

                if mapped_output_column is None or len(mapped_output_column) == 0:
                    continue
                
                entry = cast(ExcelWriter.WritableExcelFile.WritableExcelEntry, entry)

                params = ExcelWriter.WritableExcelFile.WritableExcelParameters(entry.parameters.row, entry.parameters.column)
                params.row = self.output_row_start + finalidad_index
                params.column = column_index_from_string(mapped_output_column)

                # Creating a new one because we want to show responsables once for each finalidad
                new_entry = ExcelWriter.WritableExcelFile.WritableExcelEntry(entry.content, params)

                # setattr(self.responsables, name, cell)
                entries.append(new_entry)
        
        for index, finalidad in enumerate(self.finalidades):
                        
            for name, cell_or_list_of_cells in vars(finalidad).items():
                name = cast(ColumnKey, name)
                mapped_output_column = self.output_mapping[name]

                if mapped_output_column is None or len(mapped_output_column) == 0:
                    continue
            
                if isinstance(cell_or_list_of_cells,ExcelWriter.WritableExcelFile.WritableExcelEntry):
                    entry = cast(ExcelWriter.WritableExcelFile.WritableExcelEntry, cell_or_list_of_cells)

                    entry.parameters.row = index + 1
                    entry.parameters.column = column_index_from_string(mapped_output_column)
                    # setattr(self.responsables, name, cell)
                    entries.append(entry)
                else: # it's a list of cells
                    cells = cast(list[ExcelWriter.WritableExcelFile.WritableExcelEntry], cell_or_list_of_cells)

                    if len(cells) > 0:
                        joint_cell = cells[0]
                        
                        joint_cell.content = "\n".join(
                            str(c.content) for c in cells if c.content is not None
                        )

                        joint_cell.parameters.row = index + 1
                        joint_cell.parameters.column = column_index_from_string(mapped_output_column)
                        entries.append(joint_cell)
                                        
        assert self.target is not None

        file = ExcelWriter.WritableExcelFile(self.target, entries)

        workbook = openpyxl.Workbook()
        sheet = workbook.worksheets[0]

        self._get_height_of_finalidad(sheet, 1)
        
        return file
        
        

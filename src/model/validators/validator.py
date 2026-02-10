
from collections.abc import Callable
from functools import cached_property
from io import BytesIO

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.errors import FileReadError, ValidationError
from src.model.model import FieldType, File, Model
from src.model.models.AARR_model import AARR_Model
from src.model.models.PIA_model import PIA_Model
from src.model.models.RAT_model import RAT_Model
from src.model.readers.file_model_reader import FileReader
from src.utils import get_all_file_paths_in_dir, get_file_type_from_file_path


def _validate_text_field(value: str | None) -> bool:
    """Validate if the given value is a valid text field."""
    return value is not None and value != ""

def _validate_yes_no_field(value: str | None) -> bool:
    """Validate if the given value is a valid yes/no field."""
    return value is not None and value.upper() in ["SI", "NO"]

def _validate_fixed_number_range_field(value: str | None) -> bool:
    """Validate if the given value is a valid fixed number range field."""
    return value in ["de 0 a 10.000", "+ de 100.000", "de 10.000 a 100.000"]

def _validate_treatment_duration_field(value: str | None) -> bool:
    """Validate if the given value is a valid fixed number range field."""
    return value in ["+ de 1 año", "Meses", "Instantaneo"]

def _validate_geographic_extension_field(value: str | None) -> bool:
    """Validate if the given value is a valid fixed number range field."""
    return value in ["Tratamiento a nivel nacional"]


_VALIDATORS: dict[FieldType, Callable[[str | None], bool]] = {
    "text": _validate_text_field,
    "yes_no": _validate_yes_no_field,
    "fixed_number_range": _validate_fixed_number_range_field,
    "treatment_duration": _validate_treatment_duration_field,
    "geographic_extension": _validate_geographic_extension_field
}


class Validator:

    __current_sheet: Worksheet | None
    __file_reader: FileReader

    def set_current_sheet(self, sheet: Worksheet) -> None:
        """Set the current sheet for validation."""
        self.__current_sheet = sheet

    def __validate_field_type(self, field_type: FieldType, value) -> bool | ValueError:
        """Validate the given value based on the field type."""
        assert self.__current_sheet is not None, "Current sheet is not set."

        validator = _VALIDATORS.get(field_type)
        
        if not validator:
            return ValueError(f"Unknown field type: {field_type}")
        
        return validator(value)
    
    @cached_property
    def __first_filled_row(self) -> int:
        """Find the first filled row in the given column."""
        assert self.__current_sheet is not None, "Current sheet is not set."

        max_row = self.__current_sheet.max_row

        for column_index in range(1, self.__current_sheet.max_column + 1):
            for row in range(1, max_row):
                cell_value = self.__current_sheet.cell(row=row, column=column_index).value
                if cell_value is not None and cell_value != "":
                    return row
        


        return 0
    
    @cached_property
    def __last_filled_row(self) -> int:
        """Find the last filled row in the given column."""
        assert self.__current_sheet is not None, "Current sheet is not set."

        return self.__current_sheet.max_row
    
    def __find_column_index_by_name(self, name: str) -> int | None:
        """Find the column index by its name."""
        assert self.__current_sheet is not None, "Current sheet is not set."



        # find last filled cell in row
        header_row = self.__first_filled_row
        max_column = self.__current_sheet.max_column
        for col in range(1, max_column + 1):
            cell_value = self.__current_sheet.cell(row=header_row, column=col).value
            if cell_value == name:
                return col
            
        return None
    

    
    def __get_filled_columns(self) -> list[int]:
        """Get the number of filled columns in the current sheet."""
        assert self.__current_sheet is not None, "Current sheet is not set."

        max_column = self.__current_sheet.max_column
        filled_columns = []

        for col in range(1, max_column + 1):
            if any(self.__current_sheet.cell(row=row, column=col).value is not None for row in range(1, self.__current_sheet.max_row + 1)):
                filled_columns.append(col)
        return filled_columns

    def __check_missing_model_fields_in_sheet(self, model: Model) -> list[str]:
        """Check if all fields defined in the model are present in the sheet."""
        assert self.__current_sheet is not None, "Current sheet is not set."

        missing_fields = []
        for field in model.fields:
            cell: Cell = self.__current_sheet[field.coords]

            if cell.value != field.name:
                missing_fields.append(f"Expected: {field.name}. Got: {cell.value}.")

        
        return missing_fields
    
 

    def __validate_excel_file(self, file_path: str) -> list[str] | ValidationError:
        """Validate if the given file is a valid excel file."""
        # Implement excel file validation logic here

        model: Model
        file: File

        if file_path.endswith("evaluaciones_objetivas_master.xlsx"):
            model = AARR_Model()
        elif file_path.endswith("RAT_master.xlsx"):
            model = RAT_Model()
        elif file_path.endswith("pia_master.xlsx"):
            model = PIA_Model()
        else:
            return ValidationError(message="No model found for the given file.")

        self.__file_reader = FileReader()
        self.__file_reader.set_target(file_path)
        
        read_result = self.__file_reader.read()
        
        if isinstance(read_result, FileReadError):
            print(f"Error reading file: {read_result.message}")
            return ValidationError(message=read_result.message)
        
        # parsing according to file received
        _, file = read_result # Now it's safe to assume it didn't error out, so we can get the items out of the tuple
            

        
        self.__current_sheet = openpyxl.load_workbook(BytesIO(file.data)).worksheets[0]

        validation_errors = []

        # filled_columns = self.__get_filled_columns()


        missing_field_errors = self.__check_missing_model_fields_in_sheet(model)
    
        for missing_field in missing_field_errors:
            validation_errors.append(f"Missing field '{missing_field}' in the sheet.")

        for field in model.fields:
            if field.name in missing_field_errors:
                continue

            cell: Cell = self.__current_sheet[field.coords]

            if cell.value is None:
                continue
            
            

            # Find last row of table
            last_row = self.__last_filled_row

            for row in range(cell.row + 1, last_row + 1):  # Assuming first row is header
                cell_value = self.__current_sheet.cell(row=row, column=cell.column).value

                if not field.is_required and (cell_value is None or str(cell_value).strip() == ""):
                    continue

                is_valid = self.__validate_field_type(field.data_type, cell_value)

                if isinstance(is_valid, ValueError):
                    return ValidationError(message=str(is_valid))
                if not is_valid:
                    validation_errors.append(f"Invalid value '{cell_value}' for field '{field.name}' at row {row}.")
        
        for index, error in enumerate(validation_errors):
            validation_errors[index] = f"Error in '{file.filename}': {error}"

        return validation_errors


    def validate_output(self):
        """Validate the given value based on the field type."""

        files = get_all_file_paths_in_dir("output/")

        errors = []

        for file in files:
            
            file_type = get_file_type_from_file_path(file)

            if file_type == "ExcelFile":
                errors_in_file = self.__validate_excel_file(file)
                if isinstance(errors_in_file, ValidationError):
                    errors.append(f"Error validating file '{file}': {errors_in_file.message}")
                else:
                    errors.extend(errors_in_file)

        for error in errors:
            print(error)
            
            

        
        